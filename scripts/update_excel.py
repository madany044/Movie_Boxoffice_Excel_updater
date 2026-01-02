print("🚀 Update started")

import json, re, os, shutil
import pandas as pd
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill

INPUT_PATH = "data/latest_update.txt"
EXCEL_PATH = "excel/Movie_Analytics.xlsx"
BACKUP_DIR = "excel/backups"

# ---------- SAFETY ----------
os.makedirs("excel", exist_ok=True)
os.makedirs(BACKUP_DIR, exist_ok=True)

# ---------- READ INPUT ----------
with open(INPUT_PATH, "r", encoding="utf-8") as f:
    raw = f.read()

blocks = re.split(r"---\s*(.*?)\s*---", raw)

rows = []

for i in range(1, len(blocks), 2):
    city_header = blocks[i].strip()
    json_text = blocks[i + 1].strip()

    if not json_text:
        continue

    data = json.loads(json_text)

    for show in data["shows"]:
        rows.append([
            show.get("city", city_header),
            show["venue"],
            show["time"],
            show["total"],
            show["sold"],
            show["available"],
            round(float(show["occupancy"]), 2),
            show["gross"]
        ])

df = pd.DataFrame(rows, columns=[
    "City",
    "Venue",
    "Time",
    "Total Seats",
    "Sold Seats",
    "Available Seats",
    "Occupancy %",
    "Gross"
])

print(f"📊 Total records: {len(df)}")

# ---------- BACKUP ----------
if os.path.exists(EXCEL_PATH):
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = f"{BACKUP_DIR}/Movie_Analytics_{timestamp}.xlsx"
    shutil.copy(EXCEL_PATH, backup_path)
    print("📦 Backup created")

# ---------- CREATE / LOAD EXCEL ----------
try:
    if not os.path.exists(EXCEL_PATH):
        wb = Workbook()
        wb.remove(wb.active)
        wb.save(EXCEL_PATH)

    wb = load_workbook(EXCEL_PATH)

except PermissionError:
    print("❌ ERROR: Please CLOSE Excel file and run again")
    exit()

# ---------- REMOVE OLD SHEETS ----------
for sheet in wb.sheetnames:
    del wb[sheet]

# ---------- CITY-WISE SHEETS ----------
for city in df["City"].unique():
    city_df = df[df["City"] == city]

    sheet = wb.create_sheet(title=city[:31])
    sheet.append(city_df.columns.tolist())

    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for row in city_df.itertuples(index=False):
        sheet.append(row)

# ---------- SUMMARY SHEET ----------
summary = wb.create_sheet("Summary")

summary.append([
    "City", "Total Venues", "Total Shows",
    "Total Seats", "Sold Seats",
    "Available Seats", "Total Gross",
    "Avg Occupancy %"
])

for cell in summary[1]:
    cell.font = Font(bold=True)

for city in df["City"].unique():
    city_df = df[df["City"] == city]

    summary.append([
        city,
        city_df["Venue"].nunique(),
        len(city_df),
        city_df["Total Seats"].sum(),
        city_df["Sold Seats"].sum(),
        city_df["Available Seats"].sum(),
        city_df["Gross"].sum(),
        round(city_df["Occupancy %"].mean(), 2)
    ])

# Empty row
summary.append([])

# ---------- GRAND TOTAL ----------
summary.append([
    "GRAND TOTAL",
    "",
    "",
    df["Total Seats"].sum(),
    df["Sold Seats"].sum(),
    df["Available Seats"].sum(),
    df["Gross"].sum(),
    round(df["Occupancy %"].mean(), 2)
])

last_row = summary.max_row
for cell in summary[last_row]:
    cell.font = Font(bold=True)
    cell.fill = PatternFill(
        start_color="FFF4B084",
        end_color="FFF4B084",
        fill_type="solid"
    )

# ---------- CHARTS ----------
charts = wb.create_sheet("Charts")

chart = BarChart()
chart.title = "Total Gross by City"
chart.y_axis.title = "Gross"
chart.x_axis.title = "City"

data = Reference(summary, min_col=7, min_row=1, max_row=summary.max_row - 1)
cats = Reference(summary, min_col=1, min_row=2, max_row=summary.max_row - 1)

chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)

charts.add_chart(chart, "A1")

# ---------- SAVE ----------
try:
    wb.save(EXCEL_PATH)
    print("✅ Excel updated successfully")

except PermissionError:
    print("❌ ERROR: Excel file is open. Close it and retry.")
